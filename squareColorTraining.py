# This is my first draft of an app that will allow me to practice identifying the
# square colors on a chess board. For reference, the bottom right corner of a 
# chess board is always light and the bottom left is always dark

from random import randint
from time import time, sleep
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font

# Instantiates the table
light = ('b1','d1','f1','h1','a2','c2','e2','g2','b3','d3','f3','h3','a4','c4','e4','g4',
		 'b5','d5','f5','h5','a6','c6','e6','g6','b7','d7','f7','h7','a8','c8','e8','g8')
dark = ('a1','c1','e1','g1','b2','d2','f2','h2','a3','c3','e3','g3','b4','d4','f4','h4',
		 'a5','c5','e5','g5','b6','d6','f6','h6','a7','c7','e7','g7','b8','d8','f8','h8')
board = dark + light

# Asks the user how many squares they want to practice
while True:
	run = input('How many squares do you want to train? ')
	try:
		run = int(run)
		sleep(1)
		print("Use 'l' to indicate light squares and 'd' to indicate dark squares")
		sleep(1)
		print("Begin")
		sleep(1)
		break
	except:
		print('That is not a valid amount of squares')


# Tests the user on the amount of squares they requested
start = time()
correct = 0

for i in range(run):
	square = board[randint(0,63)]		# Generates a random square
	if square in light:
		color = 'l'
	else:
		color = 'd'
	answer = input(square + ': ')		# Tests the user
	if answer == color:					# Checks for accuracy
		print('Correct!')
		correct += 1
	else:
		print('Incorrect')

# End of training session
stop = time()
duration = stop-start
accuracy = round(correct/run,3)
per_question = round(duration/run,3)
sleep(1)
print('Training complete.')
sleep(1)
print(f'Your score was {correct} out of {run} or {round(accuracy*100,1)}%')
sleep(1)
print(f'Your average speed was {per_question} seconds per square')

# Log results in an Excel file
date = datetime.now().strftime("%Y-%m-%d")
current_time = datetime.now().strftime('%H:%M:%S')
file_name = "Color Training Log V3.xlsx"

try:
    # Load existing workbook or create a new one
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Accuracy", "Reaction Time", "Total Squares", "Date", "Time"])  # Header row
        for col in range(1, 6):  # Columns A (1) to E (5)
            sheet.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=1,column=col).font = Font(bold=True)
		# Set column widths
        for col_letter in ['A','B','C','D','E']:
            sheet.column_dimensions[col_letter].width = 130 / 7.5  # Convert pixels to Excel width

    # Append the results
    sheet.append([accuracy * 100, per_question, run, date, current_time])
    row = sheet.max_row  # Get the last row where data was appended
    for col in range(1, 6):  # Columns A (1) to D (4)
        sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')

    # Save the workbook
    workbook.save(file_name)
    print(f"Results saved to {file_name}.")
except Exception as e:
    print(f"Error saving results to Excel: {e}")
